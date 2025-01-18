import csv
import time
import os
import tempfile
import smtplib
import httplib2
import platform
import subprocess
import requests
import json
import sys
import pytz
import io
import inspect
import boto3
import asyncio
import qrcode
import gc
import base64
import re

from io import BytesIO
from PIL import Image
from fpdf import FPDF
from decimal import Decimal
from typing import List
from datetime import datetime, timedelta
from dotenv import load_dotenv
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import *
from reportlab.lib.utils import *
from reportlab.graphics.barcode import code128
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from googleapiclient import discovery
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, Response, BackgroundTasks, status, Body, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import Optional
from memory_profiler import profile
from itertools import groupby
from openpyxl import Workbook, load_workbook
from PyPDF2 import PdfWriter, PdfReader
from botocore.exceptions import NoCredentialsError, PartialCredentialsError, ClientError

class UtilsController():

    def __init__(self):
        load_dotenv()

        self.REPOSITORY_NAME = self.load_env_variable('REPOSITORY_NAME')
        self._PYTHONPATH = self.load_env_variable('PYTHONPATH')

        self.SHOPIFY_API_VERSION = self.load_env_variable('SHOPIFY_API_VERSION')
        self.SHOPIFY_PRIV_KEY = self.load_env_variable('SHOPIFY_PRIV_KEY')
        self.SHOPIFY_PUB_KEY = self.load_env_variable('SHOPIFY_PUB_KEY')
        self.SHOPIFY_SECRET= self.load_env_variable('SHOPIFY_SECRET')


        self.AMAZON_ACCESS_KEY = self.load_env_variable('AMAZON_ACCESS_KEY')
        self.AMAZON_ACCESS_SECRET = self.load_env_variable('AMAZON_ACCESS_SECRET')
        self.AMAZON_REGION = self.load_env_variable('AMAZON_REGION')
        self.AMAZON_BUCKET_NAME = self.load_env_variable('AMAZON_BUCKET_NAME')
        self.AMAZON_CDN_URL = self.load_env_variable('AMAZON_CDN_URL')

        self.MACHOOL_PRIV_KEY = self.load_env_variable('MACHOOL_PRIV_KEY')
        self.MACHOOL_PUB_KEY = self.load_env_variable('MACHOOL_PUB_KEY')
        self.FARETRADE_API_KEY = self.load_env_variable('FARETRADE_API_KEY')
        self.CANADA_POST_AUTHORIZATION = self.load_env_variable('CANADA_POST_AUTHORIZATION')

        self.EMAIL_ADDRESS = self.load_env_variable('EMAIL_ADDRESS')
        self.EMAIL_PASSWORD = self.load_env_variable('EMAIL_PASSWORD')

        self.GOOGLE_SHEET_DEV_KEY = self.load_env_variable('GOOGLE_SHEET_DEV_KEY')
        self.GOOGLE_SHEET_DEV_KEY_2 = self.load_env_variable('GOOGLE_SHEET_DEV_KEY_2')
        self.SPREAD_SHEET_ID = self.load_env_variable('SPREAD_SHEET_ID')
        self.SPREAD_SHEET_ID_2 = self.load_env_variable('SPREAD_SHEET_ID_2')

        self.CURRENCYFREAKS_API_KEY = self.load_env_variable('CURRENCYFREAKS_API_KEY')

        self.RUNCOUNTER = self.load_env_variable('ORDER_RUN_COUNTER')
        self.MAXRUNS = self.load_env_variable('ORDER_MAX_RUNS')
        self.FIELDS = self.load_env_variable('ORDER_FIELDS')
        self.LIMIT = self.load_env_variable('ORDER_LIMIT')
        self.STATUS = self.load_env_variable('ORDER_STATUS')
        self.SINCE_ID = self.load_env_variable('ORDER_SINCE_ID')
        self.FULFILLMENT_STATUS = self.load_env_variable('ORDER_FULFILLMENT_STATUS')
        self.CREATED_AT_MIN = self.load_env_variable('ORDER_CREATED_AT_MIN')
        self.CREATED_AT_MAX = self.load_env_variable('ORDER_CREATED_AT_MAX')
        self.PROCESSED_AT_MIN = self.load_env_variable('ORDER_PROCESSED_AT_MIN')
        self.PROCESSED_AT_MAX = self.load_env_variable('ORDER_PROCESSED_AT_MAX')
        self.REMOVE_ORDERS = self.load_env_variable('REMOVE_ORDERS')
        self.REMOVE_LINE_ITEMS = self.load_env_variable('REMOVE_LINE_ITEMS')
        self.WAIT_TIME = self.load_env_variable('STAGING_WAIT_TIME')
        self.PHONE_CASE_LIMIT = self.load_env_variable('PHONE_CASE_LIMIT')

        self.MENU_OPTION = self.load_env_variable('MENU_OPTION')

        self.CUSTOMS_SHEET_TOKEN = self.load_env_variable('CUSTOMS_SHEET_TOKEN')
        self.CUSTOMS_SHEET_CLIENT_ID = self.load_env_variable('CUSTOMS_SHEET_CLIENT_ID')
        self.CUSTOMS_SHEET_CLIENT_SECRET = self.load_env_variable('CUSTOMS_SHEET_CLIENT_SECRET')
        self.CUSTOMS_SHEET_RANGE_NAME = self.load_env_variable('CUSTOMS_SHEET_RANGE_NAME')
        self.CUSTOMS_SPREADSHEET_ID = self.load_env_variable('CUSTOMS_SPREADSHEET_ID')
        self.DEBUG_MODE = self.load_env_variable('DEBUG_MODE')

        self.columns_array = []
        self.values_array = []
        self.condition = ""
        self.start_time = 0
        self.end_time = 0

    def get_total_time_hms(self, start_time, end_time):
        date_format = "%Y-%m-%d %H:%M:%S"
        end_time = datetime.datetime.strptime(end_time, date_format)
        start_time = datetime.datetime.strptime(start_time, date_format)
        time = end_time - start_time
        hours = str(time).split(":")[0]
        minutes = str(time).split(":")[1]
        seconds = str(time).split(":")[2]
        return hours, minutes, seconds

    def get_error_message(self, code):
        error_response = ""
        error_messages = {
            400: "Bad Request",
            401: "Unauthorized",
            402: "Payment Required",
            403: "Forbidden",
            404: "Not Found",
            405: "Method Not Allowed",
            406: "Not Acceptable",
            407: "Proxy Authentication Required",
            408: "Request Timeout",
            409: "Conflict",
            410: "Gone",
            411: "Length Required",
            412: "Precondition Failed",
            413: "Request Entity Too Large",
            414: "Request-URI Too Long",
            415: "Unsupported Media Type",
            416: "Requested Range Not Satisfiable",
            417: "Expectation Failed",
            418: "I\'m a teapot",
            419: "Page Expired (Laravel Framework)",
            420: "Method Failure (Spring Framework)",
            421: "Misdirected Request",
            422: "Unprocessable Entity",
            423: "Locked",
            424: "Failed Dependency",
            426: "Upgrade Required",
            428: "Precondition Required",
            429: "Too Many Requests",
            431: "Request Header Fields Too Large",
            440: "Login Time-out",
            444: "Connection Closed Without Response",
            449: "Retry With",
            450: "Blocked by Windows Parental Controls",
            451: "Unavailable For Legal Reasons",
            494: "Request Header Too Large",
            495: "SSL Certificate Error",
            496: "SSL Certificate Required",
            497: "HTTP Request Sent to HTTPS Port",
            498: "Invalid Token (Esri)",
            499: "Client Closed Request",
            500: "Internal Server Error",
            501: "Not Implemented",
            502: "Bad Gateway",
            503: "Service Unavailable",
            504: "Gateway Timeout",
            505: "HTTP Version Not Supported",
            506: "Variant Also Negotiates",
            507: "Insufficient Storage",
            508: "Loop Detected",
            509: "Bandwidth Limit Exceeded",
            510: "Not Extended",
            511: "Network Authentication Required",
            520: "Unknown Error",
            521: "Web Server Is Down",
            522: "Connection Timed Out",
            523: "Origin Is Unreachable",
            524: "A Timeout Occurred",
            525: "SSL Handshake Failed",
            526: "Invalid SSL Certificate",
            527: "Railgun Listener to Origin Error",
            530: "Origin DNS Error",
            598: "Network Read Timeout Error"
        }

        if code in error_messages:
            error_response += error_messages[code]
            return error_response
        else:
            return False

    def get_error_severity(self, code):
        error_response = ""
        error_messages = {
            0: "Emergency",
            1: "Alert",
            2: "Critical",
            3: "Error",
            4: "Warning",
            5: "Notification",
            6: "Informational",
            7: "Debugging"
        }

        if code in error_messages:
            error_response += error_messages[code]
            return error_response
        else:
            return False

    def replace_special_chars(self, value):
        replacements = {
            "'": "''",
            "\t": "",
            "\n": "",
            "\\t": "",
            "\\\t": "",
            "\\n": "",
            "\\\n": "",
            "\r": "", # BUG FIX - Order Examples: 4793006489677
            "\\r": "", # BUG FIX - Order Examples: 4793006489677
            "\\\r": "", # BUG FIX - Order Examples: 4793006489677
            "\\\\": "-", # BUG FIX - Order Examples: 4746625318989, 4775205699661
            '\"\\\\\"': '""',
            '\\\"': ""
        }

        for i in range(1):
            if value is not None:
                try:
                    for old, new in replacements.items():
                        value = value.replace(old, new)
                except Exception as e:
                    pass

        return value

    def send_email(self, email_from, email_to, email_subject, email_body, file_names, file_path):
        print("\n[INFO] BEGIN - Sending Email...")

        print("[INFO] Instantiating Variables...")
        # Variables
        sender_email = self.get_EMAIL_ADDRESS()
        sender_password = self.get_EMAIL_PASSWORD()
        message = MIMEMultipart()
        message["From"] = email_from
        message["To"] = ', '.join(email_to)
        message["Subject"] = email_subject
        message.attach(MIMEText(email_body, "plain"))
        attachment = None
        file_name = None
        part = None
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        server = None

        try:
            print("[INFO] Checking if it has files to be sent...")
            if file_names is not None and file_names != []:
                print("[INFO] Attaching files to be sent...")
                if file_path is not None:
                    for file_name in file_names:
                        attachment = open(file_path + file_name, 'rb')
                        print(f"[INFO] File attached... File Name: {file_name}")
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header('Content-Disposition', f'attachment; filename = {file_name}')
                        message.attach(part)
                else:
                    print("[ERROR] File path is None.")

            print("[INFO] Sending email...")
            try:
                server = smtplib.SMTP(smtp_server, smtp_port)
                server.starttls()
                server.login(sender_email, sender_password)
                server.sendmail(sender_email, email_to, message.as_string())
                print("[INFO] Email Sent...")
                return True, "[DONE] Email Sent..."
            except Exception as e:
                print(f"[ERROR] Error while Sending Email: {e}")
                return False, f"[ERROR] Error while Sending Email: {e}"
        except Exception as e:
            print(f"[ERROR] Error while Sending Email: {e}")
            return False, f"[ERROR] Error while Sending Email: {e}"
        finally:
            print("[INFO] Clearing Variables...")
            try:
                del sender_email
                del sender_password
                del message
                del attachment
                del file_name
                del part
                del smtp_server
                del smtp_port
                del server
            except:
                pass
            print("[INFO] END - Finished sending email...")

    def send_exception_email(self, module, function, error, additional_info, start_time, end_time):
        email_subject = f"[{module}] Error Exception - Crash Report"
        email_to = ['XXXXXXXXXX@COMPANY_NAME.com']
        email_from = "XXXXXXXXXX@gmail.com"

        email_body = f'[INFO] Error Exception on Module: {module}, funtion: {function}:'
        email_body += f'\n\n[ERROR DETAILS]'
        email_body += f'\n[INFO] Error Description: {error}'
        email_body += f'\n\n[EXECUTION DETAILS]'

        if start_time != 0 and end_time != 0:
            hours, minutes, seconds = self.get_total_time_hms(start_time=start_time, end_time=end_time)
            email_body += f'\n[INFO] Initial time: {start_time}'
            email_body += f'\n[INFO] Final time: {end_time}'
            email_body += f'\n[INFO] Total time taken: {hours} hours, {minutes} minutes, {seconds} seconds'

        email_body += f'\n\n[ADDITIONAL INFORMATION]'
        email_body += f'\n[INFO] Function: {function}'
        email_body += f'\n{additional_info}\n\n' if additional_info is not None else ''

        self.send_email(email_from=email_from, email_to=email_to, email_subject=email_subject, email_body=email_body, file_names=None, file_path=None)

    def convert_date_format(self, date_string, input_format, output_format):
        date = datetime.strptime(date_string, input_format)

        return date.strftime(output_format)

    def get_current_date(self):
        toronto_tz = pytz.timezone('America/Toronto')
        now = datetime.datetime.now(toronto_tz)
        return now.strftime("%Y-%m-%d")

    def get_current_date_time(self):
        toronto_tz = pytz.timezone('America/Toronto')
        now = datetime.datetime.now(toronto_tz)
        return now.strftime("%Y-%m-%d %H:%M:%S")

    def get_x_hours_ago(self, hours):
        toronto_tz = pytz.timezone('America/Toronto')
        now = datetime.datetime.now(toronto_tz)
        x_hours_ago = now - timedelta(hours=hours)
        return x_hours_ago.strftime("%Y-%m-%d %H:%M:%S")

    def get_x_days_ago(self, days):
        toronto_tz = pytz.timezone('America/Toronto')
        now = datetime.datetime.now(toronto_tz)
        x_days_ago = now - timedelta(days=days)
        return x_days_ago.strftime("%Y-%m-%d %H:%M:%S")

    def calculate_percentage(self, value_1, value_2):
        if value_1 == 0:
            return 0
        percentage_change = ((value_2 - value_1) / value_1) * 100
        return round(percentage_change, 2)

    def verify_is_processing(self, orderTags):
        tags = orderTags.split(',')

        is_processing = False

        for tag in tags:
            tagStrip = tag.strip()
            if tagStrip == 'processing_ship_done_canada':
                is_processing = True
                break

        return is_processing, orderTags

    def print_file(self, file_path):
        try:
            if platform.system() == 'Windows':
                os.startfile(file_path, 'print')
            elif platform.system() == 'Darwin':
                subprocess.run(['lp', file_path])
            elif platform.system() == 'Linux':
                subprocess.run(['lp', file_path])
            else:
                print("Unsupported operating system.")
        except Exception as e:
            print(f"Error printing PDF: {e}")

    def read_document(self, text_path):
        try:
            with open(text_path, 'r') as file:
                content = file.read()
        except Exception as e:
            return f"Error printing text: {e}"

        return content

    def get_invalid_adresses(self):
        invalid_address = {}

        print('[INFO] Getting invalid addresses.')
        discoveryUrl = ('https://sheets.googleapis.com/$discovery/rest?version=v4')
        service = discovery.build('sheets', 'v4', http=httplib2.Http(), discoveryServiceUrl=discoveryUrl, developerKey=self.get_GOOGLE_SHEET_DEV_KEY())

        spreadsheetId = self.get_SPREAD_SHEET_ID()
        rangeName = 'A1:A9999999'
        result = service.spreadsheets().values().get(spreadsheetId=spreadsheetId, range=rangeName).execute()
        values = result.get('values', [])

        if not values:
            return invalid_address, '[ERROR] No data found in the spreadsheet.'
        else:
            for row in values:
                for cel in row:
                    invalid_address[cel] = True
                    if cel.startswith('#'):
                        output_string = cel[1:]
                        invalid_address[output_string] = True
                    else:
                        invalid_address[('#' + cel)] = True

        return invalid_address, '[DONE] Finished getting invalid addresses.'

    def generate_image_pdf(self, output_file_path, output_file_name, page_width_cm, page_height_cm, image_width_cm, image_height_cm, images_path, image_names, space_between_images_cm):
        print("[INFO] Generating PDF...")

        file_path = os.path.join(output_file_path, output_file_name)
        c = canvas.Canvas(filename=file_path, pagesize=letter, pageCompression=None, invariant=None, verbosity=0, encrypt=None, cropMarks=None, pdfVersion=None, enforceColorSpace=None, initialFontName=None, initialFontSize=None, initialLeading=None, cropBox=None, artBox=None, trimBox=None, bleedBox=None, lang=None)
        c.setPageSize((page_width_cm * 28.3465, page_height_cm * 28.3465))  # Convert centimeters to points (1 cm = 28.3465 points)

        x_start = 0
        y_start = 0
        cumulative_width = 0
        cumulative_height = 0

        for image_name in image_names:
            image_path = os.path.join(images_path, image_name)
            
            if not os.path.exists(image_path):
                raise FileNotFoundError(f"Image file not found: {image_path}")

            if cumulative_width + image_width_cm * 28.3465 > (page_width_cm * 28.3465):
                x_start = 0
                cumulative_width = 0
                
                cumulative_height += image_height_cm * 28.3465 + space_between_images_cm * 28.3465
                y_start = cumulative_height
                
                if y_start + image_height_cm * 28.3465 + space_between_images_cm * 28.3465 > (page_height_cm * 28.3465):
                    c.showPage()
                    y_start = 0
                    cumulative_height = 0
            else:
                x_start = cumulative_width

            c.drawImage(image_path, x_start, y_start, width=image_width_cm * 28.3465, height=image_height_cm * 28.3465, preserveAspectRatio=True)
            cumulative_width += (image_width_cm * 28.3465) + (space_between_images_cm * 28.3465)

        c.save()
        return '[DONE] PDF generated.'

    def generate_phone_cases_pdf(self, batch_number, output_file_path, output_file_name, page_width_cm, page_height_cm, image_width_cm, image_height_cm, phone_cases_array, space_between_images_cm):
        print(f"\n[INFO] BEGIN - Generating Phone Case PDF(s)...")

        print("[INFO] Instantiating Variables...")
        # Variables
        order_name = None
        order_id = None
        item_name = None
        x_start = 20
        y_start = 20
        cumulative_width = 20
        cumulative_height = 20
        inches = 28.3465
        array_counter = 0
        phone_cases_array_length = len(phone_cases_array)

        try:
            print("[INFO] Creating PDF File...")
            print(f"[INFO] PDF Output File: {output_file_path + output_file_name}")
            file_path = os.path.join(output_file_path, output_file_name)
            pdf = canvas.Canvas(filename=file_path, pagesize=letter)
            pdf.setPageSize((page_width_cm * inches, page_height_cm * inches))

            print(f"[INFO] Looping through phone_cases_array... Size: {str(phone_cases_array_length)} items.")
            print(f"[INFO] Drawing Phone Cases on PDF...")
            for phone_case in phone_cases_array:
                array_counter += 1

                image_link = phone_case.get('phone_case_file_path')
                order_name = phone_case.get('order_name')
                # order_id = phone_case.get('order_id')
                item_name = phone_case.get('item_name')

                barcode39 = code128.Code128(order_name, barHeight = 1 * inches, barWidth=1.5)
                codes = [barcode39]

                # Download image from the link
                response = requests.get(image_link)
                if response.status_code == 200:
                    with tempfile.NamedTemporaryFile(delete=True) as temp_file:
                        temp_file.write(response.content)
                        temp_file.flush()

                        # Calculate position and size for the image
                        if cumulative_width + image_width_cm * inches > (page_width_cm * inches):
                            x_start = 20
                            cumulative_width = 20
                            cumulative_height += image_height_cm * inches + space_between_images_cm * inches
                            y_start = cumulative_height

                            if y_start + image_height_cm * inches + space_between_images_cm * inches > (page_height_cm * inches):
                                pdf.showPage()
                                y_start = 20
                                cumulative_height = 20
                        else:
                            x_start = cumulative_width

                        # Draw the barcode on the pdf
                        for code in codes:
                            code.drawOn(pdf, x_start + 20, 1.6 * inches) # Horizontal position, vertical position

                        pdf.setFontSize(10)

                        # Draw the Order Name on the pdf
                        pdf.saveState()
                        pdf.translate(x_start + (8.50 * inches), 1.1 * inches)
                        pdf.scale(-1, 1)
                        pdf.drawString(0, 0, str(order_name))
                        pdf.restoreState()

                        # Draw the Item Name on the pdf
                        pdf.saveState()
                        pdf.translate(x_start + (8.50 * inches), 0.6 * inches)
                        pdf.scale(-1, 1)
                        pdf.drawString(0, 0, str(item_name))
                        pdf.restoreState()

                        if batch_number != None and batch_number != "":
                            # Draw the Item BATCH Number on the pdf
                            pdf.saveState()
                            pdf.translate(x_start + (4 * inches), 1.1 * inches)
                            pdf.scale(-1, 1)
                            pdf.drawString(0, 0, f'BATCH {batch_number}')
                            pdf.restoreState()

                        # Draw the image on the canvas
                        pdf.drawImage(temp_file.name, x_start, y_start + (2.2 * inches), width=image_width_cm * inches, height=image_height_cm * inches, preserveAspectRatio=True)
                        cumulative_width += (image_width_cm * inches) + (space_between_images_cm * inches)
                        print(f"[INFO] {str(array_counter)}...{str(phone_cases_array_length)} Phone Case(s) Drawed... Order: {order_name} - Item: {item_name}")
                else:
                    print(f"Failed to download image from {image_link}")

            print("[INFO] Saving PDF...")
            pdf.save()
            return True, output_file_path + output_file_name
        except Exception as e:
            print(f"[ERROR] Error generating PDF: {e}")
            return False, f"[ERROR] Error generating PDF: {e}"
        finally:
            print("[INFO] Clearing variables...")
            try:
                del file_path
                del order_name
                del order_id
                del item_name
                del x_start
                del y_start
                del cumulative_width
                del cumulative_height
                del inches
                del pdf
            except:
                pass

            print(f"[INFO] END - Phone Case PDF(s) generated.")

    def convert_json_to_object(self, json_string):
        return json.loads(json_string)

    def convert_object_to_json(self, json_string, **kwargs):
        return json.dumps(obj=json_string, ensure_ascii=False)

    def generate_qr_code(self, output_file_path=None, data=None, file_type="PNG", box_size=10, border=4, resolution=None, fit=True, fill_color="black", back_color="white"):
        """
        Generates a QR code based on the provided data and saves it in the desired file format.

        Args:
            output_file_path (str): Directory path where the QR code will be saved.
            data (str): The data to encode into the QR code.
            file_type (str): The desired file type (e.g., PNG, JPG, PDF, HTML).
            box_size (int): Size of each square in the QR code grid.
            border (int): Border size (minimum is 4).
            resolution (tuple): Desired resolution (width, height) in pixels for the image.
            fit (bool): Automatically adjust the QR code version to fit the data.
            fill_color (str): Color of the QR code.
            back_color (str): Background color of the QR code.

        Returns:
            tuple: (bool, str) indicating success and the file path or error message.
        """
        qr = None
        img = None
        img_path = None

        if output_file_path is None or data is None:
            return False, "[ERROR] Missing parameters: output_file_path or data is None."
        else:
            try:
                # Create a QRCode instance with automatic version selection
                qr = qrcode.QRCode(
                    version=None,  # Automatically adjusts the version
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=box_size,
                    border=border,
                )
                qr.add_data(data)
                qr.make(fit=fit)  # Automatically determine the smallest grid size

                # Generate the QR code image
                img = qr.make_image(fill_color=fill_color, back_color=back_color)

                if resolution:
                    img = img.resize(resolution)

                # Sanitize the data to create a valid filename
                # print(f"[INFO] Data Name: {data}")
                sanitized_name = re.sub(r'[<>:"/\\|?*]', '_', data)
                # print(f"[INFO] Sanitized Name: {sanitized_name}")
                file_name = f"{sanitized_name}.{file_type.lower()}"
                # print(f"[INFO] File Name: {file_name}")
                img_path = os.path.join(output_file_path, file_name)
                # print(f"[INFO] Image Path Name: {img_path}")

                # # Define the output file path based on the chosen file type
                # file_name = f"{data}.{file_type.lower()}"
                # img_path = os.path.join(output_file_path, file_name)

                # Save the image in the specified file type
                if file_type.upper() == "PNG":
                    img.save(img_path)
                elif file_type.upper() == "JPG" or file_type.upper() == "JPEG":
                    img.convert("RGB").save(img_path, format="JPEG")
                elif file_type.upper() == "PDF":
                    img.convert("RGB").save(img_path, format="PDF")
                elif file_type.upper() == "HTML":
                    # Example: Embed the QR code image in a simple HTML file
                    html_content = f"<html><body><img src='data:image/png;base64,{self.img_to_base64(img)}'></body></html>"
                    with open(img_path, "w") as html_file:
                        html_file.write(html_content)
                else:
                    return False, f"[ERROR] Unsupported file type: {file_type}."

                return True, img_path
            except Exception as e:
                return False, f"[ERROR] Error generating QR Code: {e}"
            finally:
                try:
                    del qr, img, img_path
                except:
                    pass

    def img_to_base64(self, img):
        """Converts a QR code image to a Base64 string for embedding in HTML."""
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)
        img_str = base64.b64encode(buffer.read()).decode("utf-8")
        return img_str

    def generate_barcode(self, data=None, barHeight=60, barWidth=1):
        barcode = None
        codes = None

        if data is None or data == "":
            return False, "[ERROR] Missing parameter: data is None."

        try:
            barcode = code128.Code128(data, barHeight=barHeight, barWidth=barWidth)
            codes = [barcode]
            return True, codes
        except Exception as e:
            return False, f"[ERROR] Error generating Barcode: {e}"
        finally:
            try:
                del barcode, codes
            except:
                pass

    def generate_slips(self, orders_array, type, output_file_path, output_file_name, batch_number, page_width_cm, page_height_cm):
        print(f"\n[INFO] BEGIN - Generating Slip PDF(s)...")

        print("[INFO] Instantiating Variables...")
        # Variables
        inches = 28.3465
        file_path = os.path.join(output_file_path, output_file_name)
        order_count = 0
        item_count = 0
        type_slip = "PHONE CASE" if type == "phone_case" else "OTHER"
        y_position = 0
        orders_length = len(orders_array)
        total_items = 0
        current_date = time.strftime("%Y-%m-%d", time.localtime(time.time()))

        try:
            # Create a temporary PDF file to hold all the slips and shipping labels
            output = PdfWriter()

            print(f"[INFO] Looping through orders_array... Size: {str(len(orders_array))} orders.")
            if len(orders_array) > 0:
                for order in orders_array:
                    order_count += 1
                    item_count = 0
                    order_number = order.get('order_name')
                    country_code = order.get('country_code')
                    order_date = order.get('created_at')
                    line_items = order.get('line_items')
                    customer_name = order.get('customer_name')
                    shipping_label = order.get('shipping_label_url')

                    # Create a slip PDF for the current order
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf_file:
                        temp_pdf_path = temp_pdf_file.name
                        pdf = canvas.Canvas(filename=temp_pdf_path, pagesize=letter)
                        pdf.setPageSize((page_width_cm * inches, page_height_cm * inches))

                        # LOGO
                        pdf.setFont("Helvetica-Bold", 16)
                        pdf.drawString(x=3*inches, y=14*inches, text=f'COMPANY NAME')
                        pdf.line(x1=0*inches, y1=13.5*inches, x2=10.16*inches, y2=13.5*inches) # horizontal line

                        # BARCODE and ORDER NUMBER
                        barcode = code128.Code128(order_number, barHeight=60, barWidth=1)
                        barcode.drawOn(pdf, x=0.1*inches, y=11*inches)
                        pdf.drawString(x=1.15*inches, y=10.5*inches, text=f'{order_number}')

                        # LINES
                        pdf.line(x1=5*inches, y1=13.5*inches, x2=5*inches, y2=10*inches) # Vertical line
                        pdf.line(x1=5*inches, y1=12.3*inches, x2=page_width_cm*inches, y2=12.3*inches) # horizontal line
                        pdf.line(x1=5*inches, y1=11.2*inches, x2=page_width_cm*inches, y2=11.2*inches) # horizontal line
                        pdf.line(x1=0*inches, y1=10*inches, x2=10.16*inches, y2=10*inches) # horizontal line

                        pdf.setFont("Helvetica", 10)
                        pdf.drawString(x=5.15*inches, y=13.1*inches, text='COUNTRY CODE:')
                        pdf.drawString(x=5.15*inches, y=11.9*inches, text='SENT ON:')
                        pdf.drawString(x=5.15*inches, y=10.8*inches, text='SLIP NUMBER:')

                        pdf.setFont("Helvetica-Bold", 20)
                        pdf.drawString(x=5.15*inches, y=12.4*inches, text=f'{country_code}')
                        pdf.setFont("Helvetica", 16)
                        pdf.drawString(x=5.15*inches, y=11.3*inches, text=f'{current_date}')
                        pdf.drawString(x=5.15*inches, y=10.2*inches, text=f'#{order_count}')

                        # Batch information
                        pdf.setFont("Helvetica-Bold", 24)
                        pdf.drawString(x=2.2*inches, y=9*inches, text=type_slip.upper())
                        if batch_number != None and batch_number != "":
                            pdf.drawString(x=2.5*inches, y=8*inches, text=f'BATCH #{batch_number}')
                        pdf.line(x1=0*inches, y1=7.6*inches, x2=10.16*inches, y2=7.6*inches)

                        # Order Info
                        pdf.setFont("Helvetica-Bold", 16)
                        pdf.drawString(x=3*inches, y=6.8*inches, text='ORDER INFO:')
                        pdf.setFont("Helvetica-Bold", 12)
                        pdf.drawString(x=1*inches, y=6.2*inches, text=f'CUSTOMER NAME: ')
                        pdf.drawString(x=1.3*inches, y=5.6*inches, text=f'CREATED AT: ')
                        pdf.setFont("Helvetica", 12)
                        pdf.drawString(x=5*inches, y=6.2*inches, text=f'{customer_name}')
                        pdf.drawString(x=4.4*inches, y=5.6*inches, text=f'{order_date}')
                        pdf.line(x1=0*inches, y1=5.3*inches, x2=10.16*inches, y2=5.3*inches)

                        pdf.setFont("Helvetica-Bold", 16)
                        pdf.drawString(x=3.2*inches, y=4.6*inches, text=f'{type_slip.upper() + "S" if len(line_items) > 1 else type_slip.upper()}:')

                        pdf.setFont('Helvetica', 12)
                        y_position = 4.1*inches
                        for item in line_items:
                            total_items += 1
                            item_count += 1
                            item_name = item.get('name')

                            pdf.drawString(x=0.1*inches, y=y_position, text=item_name)
                            y_position = y_position-0.5*inches

                            print(f"[INFO] {str(order_count)}...{str(orders_length)} Order(s) Drawed... Order: {order_number} - Item {item_count} of {len(line_items)}: {item_name}")

                        pdf.showPage()
                        pdf.save()

                        # Append the slip PDF to the main output
                        with open(temp_pdf_path, 'rb') as slip_pdf_file:
                            slip_reader = PdfReader(slip_pdf_file)
                            for page in slip_reader.pages:
                                output.add_page(page)

                        # Explicitly delete the temporary slip PDF file
                        os.remove(temp_pdf_path)

                    # Append the shipping label PDF to the main output if available
                    if shipping_label:
                        response = requests.get(shipping_label)
                        if response.status_code == 200:
                            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_label_file:
                                temp_label_path = temp_label_file.name
                                temp_label_file.write(response.content)
                                temp_label_file.flush()
                                temp_label_file.seek(0)
                                label_reader = PdfReader(temp_label_path)
                                for page in label_reader.pages:
                                    output.add_page(page)

                            # Explicitly delete the temporary shipping label PDF file
                            os.remove(temp_label_path)
                        else:
                            print(f"Failed to download shipping label from {shipping_label}")

            # Write the final output to file
            with open(file_path, 'wb') as final_pdf:
                output.write(final_pdf)

            print(f"[INFO] Total Orders: {order_count} - Total Items: {total_items}")
            return True, file_path
        except Exception as e:
            print(f"[ERROR] Error generating PDF: {e}")
            return False, f"[ERROR] Error generating PDF: {e}"
        finally:
            print("[INFO] Clearing variables...")
            try:
                del inches, file_path, order_count, item_count, type_slip, y_position, order_number, country_code, order_date, line_items, item_name, orders_length, total_items, current_date
            except:
                pass

            print(f"[INFO] END - Slip PDF(s) generated.")

    def generate_small_size_slips(self, orders_array, type, output_file_path, output_file_name, batch_number, page_width_cm, page_height_cm):
        print(f"\n[INFO] BEGIN - Generating Slip PDF(s)...")

        print("[INFO] Instantiating Variables...")
        # Variables
        inches = 28.3465
        file_path = None
        pdf = None
        order_count = 0
        item_count = 0
        type_slip = "PHONE CASE" if type == "phone_case" else "OTHER"
        y_position = 0
        order_number = None
        country_code = None
        order_date = None
        line_items = None
        item_name = None
        item_variant_name = None
        orders_length = len(orders_array)
        total_items = 0
        current_date = time.strftime("%Y-%m-%d", time.localtime(time.time()))
        qr_code_return_flag = False
        barcode_return_flag = False
        img_path = None
        codes = None

        try:
            print("[INFO] Creating PDF File...")
            print(f"[INFO] PDF Output File: {output_file_path + output_file_name}")
            file_path = os.path.join(output_file_path, output_file_name)
            pdf = canvas.Canvas(filename=file_path, pagesize=letter)
            pdf.setPageSize((page_width_cm * inches, page_height_cm * inches))

            print(f"[INFO] Looping through orders_array... Size: {str(len(orders_array))} orders.")
            if len(orders_array) > 0:
                for order in orders_array:
                    order_count += 1
                    item_count = 0
                    order_number = order.get('order_name')
                    country_code = order.get('country_code')
                    order_date = order.get('created_at')
                    line_items = order.get('line_items')

                    for item in line_items:
                        # set an image as background
                        # pdf.drawImage(f'{self.get_base_directory()}/files/images/background/phone_case_bg.jpeg', 0, 0, width=page_width_cm * inches, height=page_height_cm * inches)

                        # BARCODE and ORDER NUMBER
                        # barcode_return_flag, codes = self.generate_barcode(data=order_number, barHeight=60, barWidth=1.1)
                        # if barcode_return_flag:
                            # for code in codes:
                            #     code.drawOn(pdf, x=0.01*inches, y=1.7*inches)
                            # pdf.drawString(x=1.5*inches, y=1.25*inches, text=f'{order_number}')

                        # Generate QR Code
                        qr_code_return_flag, img_path = self.generate_qr_code(output_file_path=output_file_path, data=order_number, box_size=10, border=4, fit=True, fill_color="black", back_color="white")

                        if qr_code_return_flag:
                            # Draw the QR Code on the PDF
                            pdf.drawImage(img_path, x=1.55*inches, y=1.85*inches, width=2*inches, height=2*inches)
                            pdf.setFont("Helvetica-Bold", 9)
                            pdf.drawString(x=1.7*inches, y=1.75*inches, text=f'{order_number}')

                            # Delete the temporary QR code image
                            if os.path.exists(img_path):
                                os.remove(img_path)

                        # Batch information
                        # pdf.setFont("Helvetica", 9)
                        # pdf.drawString(x=1.5*inches, y=1.43*inches, text=type_slip.upper())
                        pdf.setFont("Helvetica-Bold", 9)
                        if batch_number != None and batch_number != "":
                            pdf.drawString(x=1.5*inches, y=1.3*inches, text=f'BATCH #{batch_number}')

                        # # Order Info
                        pdf.setFont('Helvetica-Bold', 8)
                        y_position = 0.75*inches
                        total_items += 1
                        item_count += 1
                        item_name = item.get('title') # item_name = order.get('item_name').split(" / ")[0].replace("Phone Case", "")
                        item_variant_name = item.get('variant_title') # item_variant_name = order.get('item_name').split(" / ")[1]

                        pdf.setFont("Helvetica", 8)
                        pdf.drawString(x=0.5*inches, y=y_position, text=item_name)
                        y_position = y_position-0.3*inches
                        pdf.drawString(x=0.5*inches, y=y_position, text=f"{item_variant_name}")
                        y_position = y_position-0.3*inches

                        print(f"[INFO] {str(order_count)}...{str(orders_length)} Order(s) Drawed... Order: {order_number} - Item {item_count} of {len(line_items)}: {item_name}")

                        pdf.showPage()
                print("[INFO] Saving PDF...")
                print(f"[INFO] Total Orders: {order_count} - Total Items: {total_items}")
                pdf.save()
                return True, output_file_path + output_file_name
        except Exception as e:
            print(f"[ERROR] Error generating PDF: {e}")
            return False, f"[ERROR] Error generating PDF: {e}"
        finally:
            print("[INFO] Clearing variables...")
            try:
                del inches, file_path, pdf, order_count, item_count, type_slip, y_position, order_number, country_code, order_date, item, line_items, item_name, orders_length, total_items, current_date, qr_code_return_flag, img_path, codes, barcode_return_flag
            except:
                pass

            print(f"[INFO] END - Slip PDF(s) generated.")

    def read_google_sheet(self, sheet_token, sheet_client_id, sheet_client_secret, sheet_id, range):
        print(f"\n[INFO] BEGIN - Reading Google Sheet...")
        credentials = {
            "refresh_token": sheet_token,
            "client_id": sheet_client_id,
            "client_secret": sheet_client_secret,
        }
        service = None
        values = None
        result = None
        sheet = None

        try:
            print("[INFO] Connecting to Google Sheet...")
            credentials = Credentials.from_authorized_user_info(credentials)
            service = build("sheets", "v4", credentials=credentials)
            sheet = service.spreadsheets()

            print("[INFO] Reading Google Sheet...")
            result = (
                sheet.values()
                .get(
                    spreadsheetId=sheet_id,
                    range=range
                )
                .execute()
            )
            values = result.get("values", [])
            print("[INFO] Google Sheet readed...")
            return values
        except Exception as e:
            print(f"Error updating Google Sheet: {e}")
            return False
        finally:
            print("[INFO] END - Finished reading Google Sheet.")
            print("[INFO] Clearing variables...")
            try:
                del credentials
                del service
                del values
                del result
                del sheet
            except:
                pass

    def update_google_sheet_rows(self, sheet_token, sheet_client_id, sheet_client_secret, sheet_id, range, values, value_input_option='RAW'):
        print(f"\n[INFO] BEGIN - Updating Google Sheet...")
        credentials = {
            "refresh_token": sheet_token,
            "client_id": sheet_client_id,
            "client_secret": sheet_client_secret,
        }
        service = None
        request = None
        response = None

        try:
            print("[INFO] Connecting to Google Sheet...")
            credentials = Credentials.from_authorized_user_info(credentials)
            service = build("sheets", "v4", credentials=credentials)
            print("[INFO] Updating Google Sheet...")
            request = service.spreadsheets().values().update(
                spreadsheetId=sheet_id,
                range=range,
                valueInputOption=value_input_option,
                body={'values': [values]}
            )

            response = request.execute()
            if response.get('updatedCells') > 0:
                print("[INFO] Rows updated in Google Sheet.")
                return True
            else:
                return False
        except HttpError as e:
            print('Error updating row:', e)
            return False
        finally:
            print("[INFO] END - Finished updating Google Sheet.")
            print("[INFO] Clearing variables...")
            try:
                del credentials
                del service
                del request
                del response
            except:
                pass

    def insert_google_sheet_rows(self, sheet_token, sheet_client_id, sheet_client_secret, sheet_id, range, values, value_input_option='RAW'):
        print(f"\n[INFO] BEGIN - Inserting Rows into Google Sheet...")
        credentials_info = {
            "refresh_token": sheet_token,
            "client_id": sheet_client_id,
            "client_secret": sheet_client_secret,
        }
        service = None
        request = None
        response = None

        try:
            print("[INFO] Connecting to Google Sheet...")
            credentials = Credentials.from_authorized_user_info(credentials_info)
            service = build("sheets", "v4", credentials=credentials)
            print("[INFO] Inserting rows into Google Sheet...")
            request = service.spreadsheets().values().append(
                spreadsheetId=sheet_id,
                range=range,
                valueInputOption=value_input_option,
                body={'values': values}
            )

            response = request.execute()
            if response.get('updates') and response['updates']['updatedCells'] > 0:
                print("[INFO] Rows inserted into Google Sheet.")
                return True
            else:
                return False
        except HttpError as e:
            print('Error inserting rows:', e)
            return False
        finally:
            print("[INFO] END - Finished inserting rows into Google Sheet.")
            print("[INFO] Clearing variables...")
            try:
                del credentials
                del service
                del request
                del response
            except:
                pass

    def create_excel_sheet_if_not_exists(self, file_path, file_name, sheet_name) -> Workbook:
        print(f"\n[INFO] BEGIN - Creating Excel Sheet if not exists...")

        wb = None
        file_exists = True if os.path.exists(file_path + file_name) else False

        try:
            if file_exists:
                print("[INFO] Trying to load the workbook...")
                wb = load_workbook(file_path + file_name)
                print("[INFO] Workbook loaded.")
            else:
                wb = Workbook()

            print("[INFO] Checking if sheet exists...")
            # print(wb.sheetnames)

            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            if sheet_name not in wb.sheetnames:
                print(f"[INFO] Sheet not found. Creating sheet...")
                wb.create_sheet(title=sheet_name)
                print(f"[INFO] Sheet {sheet_name} created.")
                wb.save(file_path + file_name)
                print(f"[INFO] Workbook saved.")
            return wb
        except Exception as e:
            print(f"[ERROR] Error loading workbook: {e}")
            return False
        finally:
            print("[INFO] END - Finished creating Excel Sheet if not exists.")
            print("[INFO] Clearing variables...")
            try:
                del wb
                del file_exists
            except:
                pass

    def delete_files(self, sheet_file_path, pdf_file_path, sheet_file_names, pdf_file_names):
        print(f"\n[INFO] BEGIN - Deleting files")
        try:
            if len(sheet_file_names) == 0 and len(pdf_file_names) == 0:
                print(f"[INFO] No files to delete")
                return False

            if sheet_file_names is not None:
                if len(sheet_file_names) > 0:
                    for file_name in sheet_file_names:
                        os.remove(f"{sheet_file_path}{file_name}")
                        print(f"[INFO] Deleted file {file_name}")

            if pdf_file_names is not None:
                if len(pdf_file_names) > 0:
                    for file_name in pdf_file_names:
                        os.remove(f"{pdf_file_path}{file_name}")
                        print(f"[INFO] Deleted file {file_name}")
            return True
        except Exception as e:
            print(f"[ERROR] - {e}")
            return False
        finally:
            print("[INFO] Finished deleting files")
            print("[INFO] Cleaning up variables")
            try:
                del file_path
                del sheet_file_path
                del pdf_file_path
                del sheet_file_names
                del pdf_file_names
            except:
                pass

    def read_excel_file(self, file_path, file_name, sheet_name):
        print(f"\n[INFO] BEGIN - Reading Excel File...")
        sheet_header = []
        sheet_data = []
        workbook = None
        sheet = None
        row = None
        counter = 0

        try:
            print(f"[INFO] Loading workbook... File: {file_path}{file_name}")
            workbook = load_workbook(filename=file_path+file_name)

            print(f"[INFO] Reading sheet... Sheet: {sheet_name}")
            sheet = workbook[sheet_name] if sheet_name is not None and sheet_name != "" else workbook.active

            print(f"[INFO] Iterating through rows...")
            for row in sheet.iter_rows(values_only=True):
                counter += 1
                if counter == 1:
                    print(f"[INFO] Reading Header Row...")
                    sheet_header = row
                    print(f"[INFO] Header row:\t{sheet_header}")
                    print(f"[INFO] Reading Data...")
                else:
                    row_dict = dict(zip(sheet_header, row))
                    sheet_data.append(row_dict)
                    print(f"[INFO] Data row {counter}...\t{row}")

            print(f"[INFO] Finished reading Excel File. Total rows: {counter}")

            print(f"[INFO] Closing workbook...")
            workbook.close()

            return True, sheet_data, "Success"
        except FileNotFoundError:
            print(f"File '{file_path}' not found.")
            return False, None, f"File '{file_path}' not found."
        except Exception as e:
            print(f"An error occurred: {e}")
            return False, None, f"An error occurred: {e}"
        finally:
            print("[INFO] END - Finished reading Excel File.")
            print("[INFO] Clearing variables...")
            try:
                del workbook
                del sheet
                del row
                del counter
            except:
                pass
    
    def validade_file_type(self, file: UploadFile):
        file_name = None
        file_type = None
        file_extension = None
        file_path = None
        base_directory = self.get_base_directory()
        result_flag = True
        result_message = "Valid File"
        file_type_directories = {
                ".csv": "/files/csv/",
                ".ttf": "/files/fonts/",
                ".otf": "/files/fonts/",
                ".woff": "/files/fonts/",
                ".woff2": "/files/fonts/",
                ".jpg": "/files/images/",
                ".jpeg": "/files/images/",
                ".png": "/files/images/",
                ".gif": "/files/images/",
                ".bmp": "/files/images/",
                ".tiff": "/files/images/",
                ".pdf": "/files/pdfs/"
            }
        try:
            print(f"[INFO] START - Validade File Type...")
            # Extract file name and extension
            file_name = file.filename
            file_type = file_name.upper().split('.')[-1]
            file_extension = f".{file_type.lower()}"

            # Check if the file extension is recognized and get the directory
            if file_extension in file_type_directories:
                file_path = base_directory + file_type_directories[file_extension]
            else:
                raise Exception("The uploaded file is not a recognized type")

            return result_flag, file_type, file_path, result_message

        except Exception as e:
            print(f"[ERROR] Failed to upload file '{file_name}'. Error: {str(e)}")
            result_flag = False
            result_message = str(e)
            return result_flag, file_type, file_path, result_message
        finally:
            print(f"[INFO] END - Finished Validade File Type...")
            print("\n[INFO] Clearing variables...")
            try:
                del file_name, file_type, file_extension, file_path, base_directory, result_flag, result_message, file_type_directories
                gc.collect()
            except Exception as e:
                pass

    def upload_file(self, file_path: str, file_contents: bytes) -> bool:
        try:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            with open(file_path, "wb") as f:
                f.write(file_contents)

            return True
        except Exception as e:
            print(f"Error uploading file: {e}")
            return False

    def read_file(self, file_path, file_name):
        full_file_path = os.path.join(file_path, file_name)
        try:
            with open(full_file_path, 'rb') as file:
                return file.read()
        except Exception as e:
            print(f"Error reading file: {e}")
            return False

    def rename_file(self, file_path, old_file_name, new_file_name):
        old_full_file_path = os.path.join(file_path, old_file_name)
        new_full_file_path = os.path.join(file_path, new_file_name)
        try:
            os.rename(old_full_file_path, new_full_file_path)
            return True, new_file_name
        except Exception as e:
            print(f"Failed to rename file: {e}")
            return False, old_file_name

    def read_csv_file(self, file_path, file_name):
        print(f"\n[INFO] BEGIN - Reading CSV File...")
        sheet_data = []
        counter = 0

        try:
            print(f"[INFO] Reading CSV File... File: {file_path}{file_name}")
            with open(file_path + file_name, 'r') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    counter += 1
                    sheet_data.append(row)
                    print(f"[INFO] Data row {counter}...\t{row}")

            print(f"[INFO] Finished reading CSV File. Total rows: {counter}")

            return True, sheet_data, "Success"
        except FileNotFoundError:
            print(f"[WARN] File '{file_path}{file_name}' not found.")
            return False, None, f"File '{file_path}{file_name}' not found."
        except Exception as e:
            print(f"An error occurred: {e}")
            return False, None, f"An error occurred: {e}"
        finally:
            print("[INFO] END - Finished reading CSV File.")
            print("[INFO] Clearing variables...")
            try:
                del counter
            except:
                pass

    def upload_files_to_s3(self, file_directory, file_names, bucket_name):
        print(f"\n[INFO] BEGIN - Uploading files to S3... Bucket: {bucket_name}")
        s3_client = boto3.client(
            's3',
            aws_access_key_id=self.get_AMAZON_ACCESS_KEY(),
            aws_secret_access_key=self.get_AMAZON_ACCESS_SECRET()
            # region_name=region_name
        )

        results = []

        try:
            for file_name in file_names:
                file_path = os.path.join(file_directory, file_name)
                print(f"[INFO] Uploading file: {file_path}")
                try:
                    if not os.path.isfile(file_path):
                        raise FileNotFoundError(f"File not found: {file_path}")

                    s3_client.upload_file(file_path, bucket_name, file_name)
                    file_url = f"https://{bucket_name}.s3.amazonaws.com/{file_name}"
                    print(f"[INFO] File uploaded: {file_url}")

                    results.append({
                        'file_name': file_name,
                        'success': True,
                        'url': file_url,
                        'error': None
                    })
                except FileNotFoundError as e:
                    print(f"[ERROR] {str(e)}")
                    results.append({
                        'file_name': file_name,
                        'success': False,
                        'url': None,
                        'error': str(e)
                    })
                except NoCredentialsError:
                    print("[ERROR] Credentials not available")
                    results.append({
                        'file_name': file_name,
                        'success': False,
                        'url': None,
                        'error': 'Credentials not available'
                    })
                except PartialCredentialsError as e:
                    print(f"[ERROR] {str(e)}")
                    results.append({
                        'file_name': file_name,
                        'success': False,
                        'url': None,
                        'error': str(e)
                    })
                except ClientError as e:
                    print(f"[ERROR] {e.response['Error']['Message']}")
                    results.append({
                        'file_name': file_name,
                        'success': False,
                        'url': None,
                        'error': e.response['Error']['Message']
                    })
                except Exception as e:
                    print(f"[ERROR] {str(e)}")
                    results.append({
                        'file_name': file_name,
                        'success': False,
                        'url': None,
                        'error': str(e)
                    })
            return results
        except Exception as e:
            print(f"[ERROR] Error uploading files to S3: {e}")
            return False
        finally:
            print("[INFO] END - Finished uploading files to S3.")
            print("[INFO] Clearing variables...")
            try:
                del s3_client
                del results
            except:
                pass

    def load_env_variable(self, VARIABLE_NAME):
        variable = None
        try:
            variable = None if os.getenv(VARIABLE_NAME) == "None" else os.getenv(VARIABLE_NAME)
        except:
            variable = None
        finally:
            return variable

    def register_custom_fonts(self):
        pdfmetrics.registerFont(TTFont('birds_of_paradise', f'{self.get_base_directory()}/files/fonts/birds_of_paradise.ttf'))
        pdfmetrics.registerFont(TTFont('blessed', f'{self.get_base_directory()}/files/fonts/blessed.ttf'))
        pdfmetrics.registerFont(TTFont('cooper', f'{self.get_base_directory()}/files/fonts/cooper.ttf'))
        pdfmetrics.registerFont(TTFont('exmouth', f'{self.get_base_directory()}/files/fonts/exmouth.ttf'))
        pdfmetrics.registerFont(TTFont('old_london', f'{self.get_base_directory()}/files/fonts/old_london.ttf'))
        pdfmetrics.registerFont(TTFont('roboto', f'{self.get_base_directory()}/files/fonts/roboto.ttf'))
        pdfmetrics.registerFont(TTFont('sunday_burger', f'{self.get_base_directory()}/files/fonts/sunday_burger.ttf'))

        return pdfmetrics

    def get_font_title(self, name):
        name = name.lower()
        font_title = {
            'bubble': 'Bubble',
            'block': 'Block',
            'gothic': 'Gothic',
            'script': 'Script',
            'blessed': 'Blessed',
            'exmouth': 'Exmouth',
            'rounded': 'Rounded',
            'old london': 'Old London'
        }
        if name in font_title:
            return font_title[name]
        else:
            return None

    def get_current_directory(self):
        caller_frame = inspect.stack()[1]
        caller_module = inspect.getmodule(caller_frame[0])
        if caller_module:
            return os.path.dirname(os.path.abspath(caller_module.__file__))
        else:
            return None

    def get_base_directory(self):
        return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    def get_PYTHONPATH(self):
        return self._PYTHONPATH

    def get_SHOPIFY_PRIV_KEY(self):
        return self.SHOPIFY_PRIV_KEY

    def get_SHOPIFY_PUB_KEY(self):
        return self.SHOPIFY_PUB_KEY
    
    def get_SHOPIFY_SECRET(self):
        return self.SHOPIFY_SECRET

    def get_AMAZON_ACCESS_KEY(self):
        return self.AMAZON_ACCESS_KEY

    def get_AMAZON_ACCESS_SECRET(self):
        return self.AMAZON_ACCESS_SECRET

    def get_AMAZON_REGION(self):
        return self.AMAZON_REGION

    def get_AMAZON_BUCKET_NAME(self):
        return self.AMAZON_BUCKET_NAME

    def get_AMAZON_CDN_URL(self):
        return self.AMAZON_CDN_URL

    def get_MACHOOL_PRIV_KEY(self):
        return self.MACHOOL_PRIV_KEY

    def get_MACHOOL_PUB_KEY(self):
        return self.MACHOOL_PUB_KEY

    def get_FARETRADE_API_KEY(self):
        return self.FARETRADE_API_KEY

    def get_CANADA_POST_AUTHORIZATION(self):
        return self.CANADA_POST_AUTHORIZATION

    def get_EMAIL_ADDRESS(self):
        return self.EMAIL_ADDRESS

    def get_EMAIL_PASSWORD(self):
        return self.EMAIL_PASSWORD

    def get_GOOGLE_SHEET_DEV_KEY(self):
        return self.GOOGLE_SHEET_DEV_KEY

    def get_GOOGLE_SHEET_DEV_KEY_2(self):
        return self.GOOGLE_SHEET_DEV_KEY_2

    def get_SPREAD_SHEET_ID(self):
        return self.SPREAD_SHEET_ID

    def get_SPREAD_SHEET_ID_2(self):
        return self.SPREAD_SHEET_ID_2

    def get_CURRENCYFREAKS_API_KEY(self):
        return self.CURRENCYFREAKS_API_KEY

    def get_columns_array(self):
        return self.columns_array

    def get_values_array(self):
        return self.values_array

    def return_int_or_float(self, value):
        try:
            return int(value)
        except ValueError:
            return float(value)

    def get_condition(self):
        return self.condition

    def set_condition(self, value):
        self.condition = value

    def clear_condition(self):
        self.condition = ""

    def append_columns_array(self, columns):
        self.columns_array.append(columns)

    def append_values_array(self, values):
        self.values_array.append(values)

    def append_columns_values_array(self, columns, values):
        self.append_columns_array(columns)
        self.append_values_array(values)

    def clear_columns_values_arrays(self):
        self.columns_array = []
        self.values_array = []

    def validate_columns_values(self, table_column, value):
        if value is not None and value != "":
            self.append_columns_values_array(table_column, value)
        else:
            self.append_columns_values_array(table_column, "NULL")

    def get_start_time(self):
        return self.start_time

    def get_end_time(self):
        return self.end_time

    def set_start_time(self, value):
        self.start_time = value

    def set_end_time(self, value):
        self.end_time = value

    def get_REPOSITORY_NAME(self):
        return self.REPOSITORY_NAME

    def get_RUNCOUNTER(self):
        return self.RUNCOUNTER

    def get_MAXRUNS(self):
        return self.MAXRUNS

    def get_FIELDS(self):
        return self.FIELDS

    def get_LIMIT(self):
        return self.LIMIT

    def get_STATUS(self):
        return self.STATUS

    def get_CREATED_AT_MIN(self):
        return self.CREATED_AT_MIN

    def get_CREATED_AT_MAX(self):
        return self.CREATED_AT_MAX

    def get_PROCESSED_AT_MIN(self):
        return self.PROCESSED_AT_MIN

    def get_PROCESSED_AT_MAX(self):
        return self.PROCESSED_AT_MAX

    def get_SINCE_ID(self):
        return self.SINCE_ID

    def get_FULFILLMENT_STATUS(self):
        return self.FULFILLMENT_STATUS

    def get_SHOPIFY_API_VERSION(self):
        return self.SHOPIFY_API_VERSION

    def get_REMOVE_LINE_ITEMS(self):
        return self.REMOVE_LINE_ITEMS

    def get_REMOVE_ORDERS(self):
        return self.REMOVE_ORDERS

    def get_WAIT_TIME(self):
        return self.WAIT_TIME if self.WAIT_TIME is not None else "0"

    def get_PHONE_CASE_LIMIT(self):
        return self.PHONE_CASE_LIMIT

    def get_MENU_OPTION(self):
        return self.MENU_OPTION

    def get_CUSTOMS_SHEET_TOKEN(self):
        return self.CUSTOMS_SHEET_TOKEN

    def get_CUSTOMS_SHEET_CLIENT_ID(self):
        return self.CUSTOMS_SHEET_CLIENT_ID

    def get_CUSTOMS_SHEET_CLIENT_SECRET(self):
        return self.CUSTOMS_SHEET_CLIENT_SECRET

    def get_CUSTOMS_SHEET_RANGE_NAME(self):
        return self.CUSTOMS_SHEET_RANGE_NAME

    def get_CUSTOMS_SPREADSHEET_ID(self):
        return self.CUSTOMS_SPREADSHEET_ID

    def get_DEBUG_MODE(self):
        return self.DEBUG_MODE
